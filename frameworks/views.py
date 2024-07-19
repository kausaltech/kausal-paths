from datetime import datetime

from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from io import BytesIO

from frameworks.models import FrameworkConfig
from paths.types import PathsRequest


def create_result_file(request: PathsRequest, fwc_id: int, token: str):
    fwc = get_object_or_404(FrameworkConfig, id=fwc_id)
    if token != fwc.token:
        return HttpResponseForbidden('Invalid token')

    # Create a new workbook
    wb = Workbook()
    wb.create_sheet('Results')
    wb.create_sheet('Data')
    # Create a BytesIO buffer to save the workbook
    buffer = BytesIO()
    wb.save(buffer)

    # Include the current date and time in the filename in YYYYmmdd-HHMM format
    current_time = datetime.now().strftime("%Y%m%d-%H%M")
    ic = fwc.instance_config
    fn = f'{ic.identifier}_results_{current_time}.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=%s' % fn

    # Write the workbook data to the response
    response.write(buffer.getvalue())

    return response
