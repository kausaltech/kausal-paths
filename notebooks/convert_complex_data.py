from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import polars as pl
import yaml
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from nodes.constants import VALUE_COLUMN
from notebooks.notebook_support import get_context

# Import operations from upload_new_dataset.py
from notebooks.upload_new_dataset import (
    MetricData,
    clean_dataframe,
    convert_names_to_cats,
    convert_to_standard_format,
    extract_description,
    extract_metrics,
    extract_units,
    extract_units_from_row,
    pivot_by_compound_id,
    prepare_for_dvc,
    push_to_dvc,
    to_snake_case,
)


@dataclass
class MetricSpec:
    """Defines a metric with its name, ID, unit, column name, and quantity."""

    name: str  # Display name of the metric
    id: str  # Unique identifier (snake_case)
    unit: str  # Unit of measurement
    quantity: str  # Quantity type (e.g., "energy", "emissions")
    column: str | None = None  # Column name in final dataset (defaults to id if not specified)


@dataclass
class ColumnSpec:
    """Describes what a column contains."""

    # For data columns:
    quantity: str = ''
    unit: str = ''
    year: int | None = None

    # For dimension columns (if this column contains dimension values):
    dimension_name: str = ''  # If set, this column contains values for this dimension

    # Additional dimension values (for wide format):
    dimensions: dict[str, str] = field(default_factory=dict)  # dimension_name -> value

    def is_dimension_column(self) -> bool:
        """Check if this column contains dimension values (long format)."""
        return bool(self.dimension_name)

    def is_data_column(self) -> bool:
        """Check if this column contains data values (wide format)."""
        return bool(self.quantity)


@dataclass
class DatasetSchema:
    """Generic schema for peculiar datasets with implicit structure."""

    row_identifier_name: str
    identifier_mappings: dict[str, dict[str, Any]]  # identifier -> additional attributes
    dimension_names: list[str]  # All dimension names used in this dataset
    column_specs: dict[int, ColumnSpec]  # column_number -> what it contains
    skip_rows: int = 1  # Skip explanation/header rows
    identifier_column: int = 0  # Column with row identifiers (0-based)
    has_header: bool = False  # Skip header row by default, use technical col names.
    value_column_name: str = 'Value'  # Name for the value column in output


@dataclass
class DatasetConfig:
    """Complete dataset configuration including file paths and schema."""

    input_file_path: str
    output_file_path: str | None = None  # Optional, can be set at top level
    schema: DatasetSchema | None = None
    file_type: str | None = None  # Auto-detected if None: 'csv', 'excel', or 'parquet'
    sheet_name: str | None = None  # For Excel files: specific sheet name
    sheet_year_mapping: dict[str, int] | None = None  # Map sheet names to years
    excel_range: str | None = None  # Excel range like "A3:D20"
    operations: list[dict[str, Any]] = field(default_factory=list)  # Sequence of operations to apply
    instance: str | None = None  # Instance ID for context and dimensions (used by convert_names_to_cats)
    metrics: list[MetricSpec] = field(default_factory=list)  # Metric definitions (name, id, unit)


def parse_excel_range(range_str: str) -> tuple[int, int, int, int]:
    """
    Parse Excel range like "A3:D20" into (start_row, start_col, end_row, end_col).

    Returns 0-based indices: (start_row, start_col, end_row, end_col)
    """

    # Match patterns like "A3:D20" or "A3" (single cell)
    pattern = r'([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?'
    match = re.match(pattern, range_str.upper())
    if not match:
        raise ValueError(f"Invalid Excel range format: {range_str}")

    start_col_str, start_row_str, end_col_str, end_row_str = match.groups()

    # Convert to 0-based indices
    start_col = column_index_from_string(start_col_str) - 1
    start_row = int(start_row_str) - 1

    if end_col_str and end_row_str:
        end_col = column_index_from_string(end_col_str) - 1
        end_row = int(end_row_str) - 1
    else:
        # Single cell - return same as start
        end_col = start_col
        end_row = start_row

    return start_row, start_col, end_row, end_col


class FileLoader:
    """Handles opening and reading files in different formats."""

    @staticmethod
    def detect_file_type(file_path: str | Path) -> str:
        """Detect file type from extension."""
        path = Path(file_path)
        suffix = path.suffix.lower()
        if suffix in ['.xlsx', '.xls']:
            return 'excel'
        if suffix == '.parquet':
            return 'parquet'
        return 'csv'

    @staticmethod
    def load_csv(file_path: str | Path, skip_rows: int = 0, has_header: bool = True) -> pl.DataFrame:
        """Load a CSV file into a Polars DataFrame."""
        return pl.read_csv(
            file_path,
            skip_rows=skip_rows,
            has_header=has_header,
        )

    @staticmethod
    def load_parquet(file_path: str | Path) -> pl.DataFrame:
        """Load a parquet file into a Polars DataFrame."""
        return pl.read_parquet(file_path)

    @staticmethod
    def load_excel(  # noqa: C901, PLR0912
        file_path: str | Path,
        sheet_name: str | None = None,
        skip_rows: int = 0,
        has_header: bool = False,
        excel_range: str | None = None
    ) -> pl.DataFrame:
        """Load an Excel file sheet into a Polars DataFrame."""
        wb = load_workbook(file_path, data_only=True, read_only=True)
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        if ws is None:
            wb.close()
            return pl.DataFrame()

        # Handle Excel range if specified
        if excel_range:
            start_row, start_col, end_row, end_col = parse_excel_range(excel_range)
            # Extract only the specified range
            data = []
            for row_idx in range(start_row, end_row + 1):
                row_data = []
                for col_idx in range(start_col, end_col + 1):
                    cell = ws.cell(row=row_idx + 1, column=col_idx + 1)  # openpyxl uses 1-based
                    row_data.append(cell.value)
                data.append(row_data)
        else:
            # Convert to list of lists
            data = [list(row) for row in ws.iter_rows(values_only=True)]

        wb.close()

        if not data:
            return pl.DataFrame()

        # Determine max columns
        max_cols = max(len(row) for row in data) if data else 0

        # Pad shorter rows
        padded_data = [list(row) + [None] * (max_cols - len(row)) for row in data]

        # Apply skip_rows first
        if skip_rows > 0:
            padded_data = padded_data[skip_rows:]

        if not padded_data:
            return pl.DataFrame()

        # Create column names
        if has_header and len(padded_data) > 0:
            # Use first row as header
            header_row = padded_data[0]
            col_names = [
                f"column_{i+1}" if (not h or str(h).strip() == '') else str(h).strip()
                for i, h in enumerate(header_row[:max_cols])
            ]
            data_rows = padded_data[1:]
        else:
            col_names = [f"column_{i+1}" for i in range(max_cols)]
            data_rows = padded_data

        if not data_rows:
            return pl.DataFrame()

        # Create DataFrame
        df = pl.DataFrame(data_rows, schema=col_names[:max_cols], orient="row")
        return df

    @classmethod
    def load_file(
        cls,
        file_path: str | Path,
        file_type: str | None = None,
        sheet_name: str | None = None,
        skip_rows: int = 0,
        has_header: bool = False,
        excel_range: str | None = None
    ) -> pl.DataFrame:
        """Load a file based on its type."""
        if file_type is None:
            file_type = cls.detect_file_type(file_path)

        if file_type == 'parquet':
            return cls.load_parquet(file_path)
        if file_type == 'excel':
            return cls.load_excel(file_path, sheet_name, skip_rows, has_header, excel_range)
        # Default to CSV
        return cls.load_csv(file_path, skip_rows, has_header)


class DataCollector:
    """Handles collecting and extracting data from loaded DataFrames."""

    def __init__(self, schema: DatasetSchema):
        self.schema = schema

    def collect_data(self, df: pl.DataFrame, year: int | None = None) -> list[dict[str, Any]]:  # noqa: C901, PLR0912
        """Collect data from DataFrame and return as list of row dictionaries."""
        if len(df) == 0:
            return []

        # Get column name for identifier column
        identifier_col_name = df.columns[self.schema.identifier_column]

        # Extract identifiers from specified column
        identifiers = df.select(
            pl.col(identifier_col_name).alias('identifier')
        ).to_series()

        # Transform to long format
        result_rows = []

        for row_idx, identifier in enumerate(identifiers):
            if identifier is None or str(identifier).strip() == '':
                continue

            # Get additional attributes for this identifier
            identifier_attrs = self.schema.identifier_mappings.get(str(identifier), {})

            # First, collect dimension values from dimension columns (long format)
            row_dimensions = {}
            for col_num, col_spec in self.schema.column_specs.items():
                if col_spec.is_dimension_column():
                    try:
                        dim_value = df.item(row_idx, col_num)
                        if dim_value is not None and str(dim_value).strip() != '':
                            row_dimensions[col_spec.dimension_name] = str(dim_value)
                    except (IndexError, ValueError, TypeError):
                        continue

            # Process each data column (wide format)
            for col_num, col_spec in self.schema.column_specs.items():
                if not col_spec.is_data_column():
                    continue  # Skip dimension columns

                try:
                    value = df.item(row_idx, col_num)
                    if value is not None and str(value).strip() != '':

                        # Build result row with identifier
                        row_data = {
                            self.schema.row_identifier_name: str(identifier),
                            'Quantity': col_spec.quantity,
                            'Unit': col_spec.unit,
                        }

                        # Add year if specified (from column spec or parameter)
                        if year is not None:
                            row_data['Year'] = year  # type: ignore
                        elif col_spec.year is not None:
                            row_data['Year'] = col_spec.year  # type: ignore

                        # Add dimensions from long format (from dimension columns)
                        row_data.update(row_dimensions)

                        # Add dimensions from wide format (from column spec)
                        row_data.update(col_spec.dimensions)

                        # Add value
                        row_data[self.schema.value_column_name] = float(value)  # type: ignore

                        # Add any additional attributes from identifier mapping
                        row_data.update(identifier_attrs)

                        result_rows.append(row_data)

                except (IndexError, ValueError, TypeError):
                    continue  # Skip invalid values

        return result_rows


class DataTransformer:
    """Handles converting collected data into normalized format."""

    @staticmethod
    def transform_to_dataframe(rows: list[dict[str, Any]]) -> pl.DataFrame:
        """Convert list of row dictionaries to normalized DataFrame."""
        if not rows:
            return pl.DataFrame()
        return pl.DataFrame(rows)


class OperationsExecutor:
    """Executes operations on DataFrames."""

    def __init__(self, context=None, metrics: list[MetricSpec] | None = None):
        """Initialize with optional context and metrics for operations that need them."""
        self.context = context
        self.metrics = metrics or []
        self._extracted_units: dict[str, str] = {}  # Store extracted units

    @staticmethod
    def _eval_polars_expr(expr_str: str, _df: pl.DataFrame) -> pl.Expr:
        """
        Safely evaluate a Polars expression string.

        The expression should be a Polars expression like:
        - "pl.col('value') > 0"
        - "pl.col('year') == 2020"
        - "pl.col('name').is_not_null()"
        """
        # Create a safe namespace with only polars
        namespace = {'pl': pl}

        # Compile and evaluate the expression
        try:
            code = compile(expr_str, '<string>', 'eval')
            expr = eval(code, {"__builtins__": {}}, namespace)  # noqa: S307
            if not isinstance(expr, pl.Expr):
                raise TypeError(f"Expression must evaluate to a Polars Expr, got {type(expr)}")  # noqa: TRY301
            return expr  # noqa: TRY300
        except Exception as e:
            raise ValueError(f"Invalid Polars expression '{expr_str}': {e}") from e

    def execute_operation(self, df: pl.DataFrame, operation: dict[str, Any]) -> pl.DataFrame:  # noqa: C901, PLR0911, PLR0912, PLR0915
        """Execute a single operation on a DataFrame."""
        op_type = operation.get('type')
        op_params = operation.get('params', {})

        if op_type == 'filter':
            # Filter using Polars expression
            expr_str = op_params.get('expr')
            if not expr_str:
                raise ValueError("Filter operation requires 'expr' parameter")
            expr = self._eval_polars_expr(expr_str, df)
            return df.filter(expr)

        if op_type == 'with_columns':
            # Add new columns using Polars expressions
            # Can be a single expression string or a list of expression strings
            expr_or_exprs = op_params.get('expr') or op_params.get('exprs')
            if expr_or_exprs is None:
                raise ValueError("with_columns operation requires 'expr' or 'exprs' parameter")

            if isinstance(expr_or_exprs, str):
                # Single expression
                expr = self._eval_polars_expr(expr_or_exprs, df)
                return df.with_columns(expr)
            if isinstance(expr_or_exprs, list):
                # List of expressions
                exprs = [self._eval_polars_expr(e, df) for e in expr_or_exprs]
                return df.with_columns(exprs)
            raise ValueError("with_columns 'expr'/'exprs' must be a string or list of strings")

        if op_type == 'drop':
            # Drop columns
            columns = op_params.get('columns', [])
            return df.drop(columns)

        if op_type == 'rename':
            # Rename columns
            mapping = op_params.get('mapping', {})
            return df.rename(mapping)

        if op_type == 'select':
            # Select columns
            columns = op_params.get('columns', [])
            return df.select(columns)

        # Imported operations from upload_new_dataset.py
        if op_type == 'clean_dataframe':
            return clean_dataframe(df)

        if op_type == 'convert_to_standard_format':
            return convert_to_standard_format(df)

        if op_type == 'pivot_by_compound_id':
            return pivot_by_compound_id(df)

        if op_type == 'prepare_for_dvc':
            units = op_params.get('units', {})
            return prepare_for_dvc(df, units)

        if op_type == 'to_snake_case_columns':
            # Convert column names to snake_case
            exclude = op_params.get('exclude', ['Year'])
            new_cols = {}
            for col in df.columns:
                if col not in exclude:
                    new_cols[col] = to_snake_case(col)
            return df.rename(new_cols)

        if op_type == 'convert_names_to_cats':
            # Convert dimension column names to category IDs using context
            if self.context is None:
                raise ValueError(
                    "convert_names_to_cats operation requires a context. "
                    + "Set 'instance' in the dataset configuration."
                )
            # Units must be provided explicitly via params
            units = op_params.get('units', {})
            if not isinstance(units, dict):
                raise ValueError("convert_names_to_cats operation requires 'units' to be a dict in params")
            return convert_names_to_cats(df, units, self.context)

        if op_type == 'define_metrics':
            # Create metric_col column from YAML-defined metrics
            # Maps Quantity/Metric values to metric column names based on metric definitions
            if not self.metrics:
                raise ValueError(
                    "define_metrics operation requires metrics to be defined in the dataset configuration."
                )
            # Create a mapping from metric name/quantity to metric column name
            # Use column if specified, otherwise use id
            metric_mapping = {}
            for metric in self.metrics:
                column_name = metric.column if metric.column else metric.id
                # Map by name (case-insensitive) and also by id
                metric_mapping[metric.name.lower()] = column_name
                metric_mapping[metric.id.lower()] = column_name

            # Determine which column contains metric names
            metric_col_name = None
            if 'Metric' in df.columns:
                metric_col_name = 'Metric'
            elif 'Quantity' in df.columns:
                metric_col_name = 'Quantity'
            else:
                raise ValueError(
                    "define_metrics operation requires 'Metric' or 'Quantity' column in DataFrame."
                )

            # Create metric_col by mapping metric names to column names
            def map_to_metric_column(name: str | None) -> str:
                if name is None:
                    return VALUE_COLUMN
                name_lower = str(name).lower().strip()
                return metric_mapping.get(name_lower, to_snake_case(name))

            df = df.with_columns(
                pl.col(metric_col_name)
                .map_elements(map_to_metric_column, return_dtype=pl.Utf8)
                .alias('metric_col')
            )
            return df

        if op_type == 'extract_units':
            # Extract units from DataFrame using metric_col
            # Returns units dict: {metric_id: unit}
            if 'metric_col' not in df.columns:
                raise ValueError(
                    "extract_units operation requires 'metric_col' column. "
                    + "Run 'define_metrics' operation first."
                )
            if 'Unit' not in df.columns:
                raise ValueError("extract_units operation requires 'Unit' column in DataFrame.")

            units = extract_units(df)
            # Store units in operation executor for later use
            self._extracted_units = units
            return df

        if op_type == 'extract_units_from_row':
            # Extract units from first row if it contains only strings
            units, df_cleaned = extract_units_from_row(df)
            self._extracted_units = units
            return df_cleaned

        if op_type == 'extract_metadata':
            # Separate metadata columns (Quantity, Unit, Metric) from data
            # This prepares the DataFrame for final output format
            metadata_cols = ['Quantity', 'Unit', 'Metric', 'metric_col']
            cols_to_drop = [col for col in metadata_cols if col in df.columns]
            if cols_to_drop:
                df = df.drop(cols_to_drop)
            return df

        if op_type == 'print_metadata':
            # Print metadata information for verification before pushing to DVC
            print("\n" + "=" * 80)
            print("METADATA SUMMARY")
            print("=" * 80)

            # Print units
            units_to_print = op_params.get('units')
            if units_to_print is None:
                units_to_print = self._extracted_units
            if units_to_print:
                print(f"\nUnits ({len(units_to_print)}):")
                print("-" * 80)
                for metric_id, unit in sorted(units_to_print.items()):
                    print(f"  {metric_id:30s} -> {unit}")
            else:
                print("\nUnits: None extracted")

            # Print metrics from YAML definitions
            if self.metrics:
                print(f"\nMetrics ({len(self.metrics)}):")
                print("-" * 80)
                for metric in self.metrics:
                    column_name = metric.column if metric.column else metric.id
                    print(f"  Name:     {metric.name}")
                    print(f"    ID:     {metric.id}")
                    print(f"    Column: {column_name}")
                    print(f"    Unit:   {metric.unit}")
                    print(f"    Quantity: {metric.quantity}")
                    print()
            else:
                print("\nMetrics: None defined in configuration")

            # Print DataFrame structure
            print("DataFrame Structure:")
            print("-" * 80)
            print(f"  Rows: {len(df)}")
            print(f"  Columns ({len(df.columns)}):")
            metric_cols = []
            dimension_cols = []
            other_cols = []
            units_dict = units_to_print or {}
            for col in df.columns:
                if col in units_dict:
                    metric_cols.append(col)
                elif col.lower() == 'year':
                    other_cols.append(f"{col} (time dimension)")
                else:
                    dimension_cols.append(col)
            if metric_cols:
                print(f"    Metric columns ({len(metric_cols)}): {', '.join(sorted(metric_cols))}")
            if dimension_cols:
                print(f"    Dimension columns ({len(dimension_cols)}): {', '.join(sorted(dimension_cols))}")
            if other_cols:
                print(f"    Other columns: {', '.join(other_cols)}")

            # Print sample data
            if len(df) > 0:
                sample_size = op_params.get('sample_rows', 5)
                print(f"\nSample Data (first {min(sample_size, len(df))} rows):")
                print("-" * 80)
                sample_df = df.head(sample_size)
                # Print in a readable format
                for idx, row in enumerate(sample_df.iter_rows(named=True), 1):
                    print(f"  Row {idx}:")
                    for key, value in row.items():
                        if value is not None:
                            print(f"    {key}: {value}")
                    print()

            print("=" * 80 + "\n")
            return df

        if op_type == 'push_to_dvc':
            # Push dataset to DVC repository
            output_path = op_params.get('output_path')
            if not output_path:
                raise ValueError("push_to_dvc operation requires 'output_path' parameter")

            dataset_name = op_params.get('dataset_name', 'dataset')
            language = op_params.get('language', 'en')

            # Get units (from params, YAML metric definitions, extracted units, or empty dict)
            units = op_params.get('units')
            if units is None:
                # Try to build units from YAML metric definitions first
                if self.metrics:
                    units = {}
                    for metric_spec in self.metrics:
                        column_name = metric_spec.column if metric_spec.column else metric_spec.id
                        units[column_name] = metric_spec.unit
                else:
                    # Fall back to extracted units
                    units = self._extracted_units
            if not isinstance(units, dict):
                raise ValueError("push_to_dvc operation requires 'units' to be a dict")

            # Warn if units are empty
            if not units:
                print("Warning: No units found. Units should be:")
                print("  1. Provided via params['units']")
                print("  2. Defined in YAML metrics (metric.unit)")
                print("  3. Extracted via 'extract_units' operation (before pivot/extract_metadata)")

            # Extract description if Description column exists
            description = op_params.get('description')
            if description is None and 'Description' in df.columns:
                description = extract_description(df)

            # Convert MetricSpec to MetricData for DVC
            metrics_list = []
            if self.metrics:
                # Use YAML-defined metrics
                for metric_spec in self.metrics:
                    column_name = metric_spec.column if metric_spec.column else metric_spec.id
                    metrics_list.append(MetricData(
                        id=column_name,
                        quantity=to_snake_case(metric_spec.quantity),
                        label={language: metric_spec.name}
                    ))
            elif op_params.get('extract_metrics', False):
                # Extract metrics from DataFrame if requested
                if 'metric_col' in df.columns and 'Metric' in df.columns and 'Quantity' in df.columns:
                    metrics_list = extract_metrics(df, language)
                else:
                    print("Warning: Cannot extract metrics - missing required columns (metric_col, Metric, Quantity)")

            # Push to DVC
            push_to_dvc(
                df=df,
                output_path=output_path,
                dataset_name=dataset_name,
                units=units,
                description=description,
                metrics=metrics_list,
                language=language
            )
            print(f"✓ Dataset '{dataset_name}' pushed to DVC at {output_path}")
            return df

        raise ValueError(f"Unknown operation type: {op_type}")

    def execute_operations(self, df: pl.DataFrame, operations: list[dict[str, Any]]) -> pl.DataFrame:
        """Execute a sequence of operations in order."""
        result_df = df
        for operation in operations:
            result_df = self.execute_operation(result_df, operation)
        return result_df


class OutputWriter:
    """Handles writing output files."""

    @staticmethod
    def write_csv(df: pl.DataFrame, output_path: str | Path, verbose: bool = True) -> None:
        """Write DataFrame to CSV file."""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if verbose:
            print(f"Writing output to: {output_path}")

        df.write_csv(output_path)

        if verbose:
            print(f"✓ Written {len(df)} rows to {output_path}")


class DatasetProcessor:
    """Orchestrates the complete dataset processing pipeline."""

    def __init__(self, config: DatasetConfig):
        self.config = config
        self.file_loader = FileLoader()
        self.data_collector = DataCollector(config.schema) if config.schema else None
        self.data_transformer = DataTransformer()
        self.output_writer = OutputWriter()
        # Initialize context if instance is specified
        context = None
        if config.instance:
            try:
                context = get_context(config.instance)
            except Exception as e:
                print(f"Warning: Could not load context for instance '{config.instance}': {e}")
        self.operations_executor = OperationsExecutor(context=context, metrics=config.metrics)

    def process(self, verbose: bool = True) -> pl.DataFrame:  # noqa: C901, PLR0912
        """Process a single dataset configuration."""
        if self.config.schema is None:
            # For parquet files without schema, just load and return as-is
            if verbose:
                print(f"Loading file: {self.config.input_file_path}")
            df = self.file_loader.load_file(
                self.config.input_file_path,
                self.config.file_type,
                excel_range=self.config.excel_range
            )

            # Apply operations if any
            if self.config.operations:
                df = self.operations_executor.execute_operations(df, self.config.operations)

            return df

        # Full processing pipeline
        if verbose:
            print(f"Processing: {self.config.input_file_path}")
            if self.config.file_type == 'excel':
                print("  File type: Excel")
                if self.config.sheet_name:
                    print(f"  Sheet: {self.config.sheet_name}")
                elif self.config.sheet_year_mapping:
                    print("  Processing multiple sheets with year mapping")
                if self.config.excel_range:
                    print(f"  Excel range: {self.config.excel_range}")

        if self.data_collector is None:
            raise ValueError("Schema is required for data collection")

        # Step 1: Open file(s)
        if self.config.file_type == 'excel' and self.config.sheet_year_mapping:
            # Process multiple sheets
            all_rows = []
            for sheet_name, year in self.config.sheet_year_mapping.items():
                if verbose:
                    print(f"\n  Processing sheet '{sheet_name}' (year: {year})")

                df = self.file_loader.load_file(
                    self.config.input_file_path,
                    file_type='excel',
                    sheet_name=sheet_name,
                    skip_rows=self.config.schema.skip_rows,
                    has_header=self.config.schema.has_header,
                    excel_range=self.config.excel_range
                )

                # Apply operations before collecting data
                if self.config.operations:
                    df = self.operations_executor.execute_operations(df, self.config.operations)

                # Step 2: Collect data
                rows = self.data_collector.collect_data(df, year=year)
                all_rows.extend(rows)

                if verbose:
                    print(f"    Collected {len(rows)} rows from this sheet")

            # Step 3: Transform to DataFrame
            result_df = self.data_transformer.transform_to_dataframe(all_rows)

        elif self.config.file_type == 'excel' and self.config.sheet_name:
            # Process single sheet
            df = self.file_loader.load_file(
                self.config.input_file_path,
                file_type='excel',
                sheet_name=self.config.sheet_name,
                skip_rows=self.config.schema.skip_rows,
                has_header=self.config.schema.has_header,
                excel_range=self.config.excel_range
            )

            # Apply operations before collecting data
            if self.config.operations:
                df = self.operations_executor.execute_operations(df, self.config.operations)

            rows = self.data_collector.collect_data(df)
            result_df = self.data_transformer.transform_to_dataframe(rows)

        else:
            # Process single file (CSV, parquet, or Excel default sheet)
            df = self.file_loader.load_file(
                self.config.input_file_path,
                file_type=self.config.file_type,
                skip_rows=self.config.schema.skip_rows,
                has_header=self.config.schema.has_header,
                excel_range=self.config.excel_range
            )

            # Apply operations before collecting data
            if self.config.operations:
                df = self.operations_executor.execute_operations(df, self.config.operations)

            rows = self.data_collector.collect_data(df)
            result_df = self.data_transformer.transform_to_dataframe(rows)

        return result_df

    def get_summary(self) -> dict[str, Any]:
        """Get summary of schema configuration."""
        if self.config.schema is None:
            return {}

        data_columns = [spec for spec in self.config.schema.column_specs.values() if spec.is_data_column()]
        dimension_columns = [spec for spec in self.config.schema.column_specs.values() if spec.is_dimension_column()]

        summary = {
            'total_columns': len(self.config.schema.column_specs),
            'data_columns': len(data_columns),
            'dimension_columns': len(dimension_columns),
            'dimensions': self.config.schema.dimension_names,
            'identifiers': len(self.config.schema.identifier_mappings)
        }

        if data_columns:
            summary['quantities'] = list(set(spec.quantity for spec in data_columns if spec.quantity))
            summary['units'] = list(set(spec.unit for spec in data_columns if spec.unit))
            years = [spec.year for spec in data_columns if spec.year is not None]
            if years:
                summary['years'] = sorted(set(years))  # type: ignore

        return summary


def list_excel_sheets(file_path: str | Path) -> list[str]:
    """List all sheet names in an Excel file."""
    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    if file_path.suffix.lower() not in ['.xlsx', '.xls']:
        raise ValueError(f"File is not an Excel file: {file_path}")

    wb = load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    return sheet_names


def merge_configs(base_config: dict[str, Any], override_config: dict[str, Any]) -> dict[str, Any]:
    """Merge two config dictionaries, with override taking precedence."""
    merged = base_config.copy()

    # Deep merge for nested dictionaries
    for key, value in override_config.items():
        if key in merged and isinstance(merged[key], dict) and isinstance(value, dict):
            merged[key] = merge_configs(merged[key], value)
        elif key in merged and isinstance(merged[key], list) and isinstance(value, list):
            # For lists, replace entirely (don't merge)
            merged[key] = value
        else:
            merged[key] = value

    return merged


def parse_column_specs(config_dict: dict[str, Any]) -> dict[int, ColumnSpec]:
    """Parse column specifications from config dictionary."""
    column_specs = {}
    for col_num_str, spec_dict in config_dict.get('column_specs', {}).items():
        col_num = int(col_num_str)

        # Create ColumnSpec based on whether it's a dimension or data column
        if spec_dict.get('dimension_name'):
            # This is a dimension column (long format)
            column_specs[col_num] = ColumnSpec(
                dimension_name=spec_dict['dimension_name']
            )
        else:
            # This is a data column (wide format)
            column_specs[col_num] = ColumnSpec(
                quantity=spec_dict.get('quantity', ''),
                unit=spec_dict.get('unit', ''),
                year=spec_dict.get('year'),
                dimensions=spec_dict.get('dimensions', {})
            )
    return column_specs


def parse_schema(config_dict: dict[str, Any]) -> DatasetSchema:
    """Parse dataset schema from config dictionary."""
    column_specs = parse_column_specs(config_dict)

    return DatasetSchema(
        row_identifier_name=config_dict['row_identifier_name'],
        skip_rows=config_dict.get('skip_rows', 1),
        has_header=config_dict.get('has_header', False),
        identifier_column=config_dict.get('identifier_column', 0),
        identifier_mappings=config_dict.get('identifier_mappings', {}),
        column_specs=column_specs,
        dimension_names=config_dict.get('dimension_names', []),
        value_column_name=config_dict.get('value_column_name', 'Value')
    )


def parse_metrics(config_list: list[dict[str, Any]] | None) -> list[MetricSpec]:
    """Parse metric specifications from config list."""
    if not config_list:
        return []
    return [
        MetricSpec(
            name=metric_dict['name'],
            id=metric_dict['id'],
            unit=metric_dict['unit'],
            quantity=metric_dict['quantity'],
            column=metric_dict.get('column')  # Optional, defaults to None
        )
        for metric_dict in config_list
    ]


def parse_dataset_config(
    config_dict: dict[str, Any],
    default_output: str | None = None,
    base_config: dict[str, Any] | None = None
) -> DatasetConfig:
    """Parse a single dataset configuration from dictionary."""
    # Merge with base config if provided (property inheritance)
    if base_config:
        config_dict = merge_configs(base_config, config_dict)

    # Determine file type
    input_path = Path(config_dict['input_file_path'])
    file_type = FileLoader.detect_file_type(input_path)

    # Parse schema if present
    schema = None
    if 'row_identifier_name' in config_dict:
        schema = parse_schema(config_dict)

    # Parse metrics if present
    metrics = parse_metrics(config_dict.get('metrics'))

    return DatasetConfig(
        input_file_path=config_dict['input_file_path'],
        output_file_path=config_dict.get('output_file_path', default_output),
        schema=schema,
        file_type=config_dict.get('file_type', file_type),
        sheet_name=config_dict.get('sheet_name'),
        sheet_year_mapping=config_dict.get('sheet_year_mapping'),
        excel_range=config_dict.get('excel_range'),
        operations=config_dict.get('operations', []),
        instance=config_dict.get('instance'),
        metrics=metrics
    )


def load_config(yaml_file_path: str | Path) -> tuple[list[DatasetConfig], str | None]:
    """
    Load dataset configurations from YAML file.

    Supports both single config and list of configs.
    Implements property inheritance: later datasets inherit from previous ones.

    Args:
        yaml_file_path: Path to the YAML configuration file

    Returns:
        Tuple of (list of DatasetConfig objects, output_file_path)

    """
    yaml_file_path = Path(yaml_file_path)

    if not yaml_file_path.exists():
        raise FileNotFoundError(f"Configuration file not found: {yaml_file_path}")

    with yaml_file_path.open('r', encoding='utf-8') as f:
        config = yaml.safe_load(f)

    # Handle top-level output path, instance, and metrics
    default_output: str | None = config.get('output_file_path')
    top_level_instance: str | None = config.get('instance')
    top_level_metrics: list[dict[str, Any]] | None = config.get('metrics')

    # Check if config is a list of datasets
    if 'datasets' in config:
        # Multiple dataset configurations with inheritance
        dataset_configs = []
        # Start with top-level properties as base config
        base_config: dict[str, Any] = {}
        if top_level_instance is not None:
            base_config['instance'] = top_level_instance
        if top_level_metrics is not None:
            base_config['metrics'] = top_level_metrics

        for dataset_dict in config['datasets']:
            # Each dataset inherits from the previous one (base_config)
            dataset_config = parse_dataset_config(dataset_dict, default_output, base_config)
            dataset_configs.append(dataset_config)

            # Update base_config for next iteration (current config becomes base for next)
            base_config = merge_configs(base_config, dataset_dict)

        return dataset_configs, default_output

    # Single dataset configuration (backward compatibility)
    # Include top-level instance and metrics if present
    single_base_config: dict[str, Any] = {}
    if top_level_instance is not None:
        single_base_config['instance'] = top_level_instance
    if top_level_metrics is not None:
        single_base_config['metrics'] = top_level_metrics
    dataset_config = parse_dataset_config(config, default_output, single_base_config if single_base_config else None)
    return [dataset_config], default_output


def process_datasets(config_path: str | Path, verbose: bool = True) -> pl.DataFrame:  # noqa: C901, PLR0912
    """
    Process one or more datasets using configuration from YAML file.

    Args:
        config_path: Path to YAML configuration file
        verbose: Whether to print progress information

    Returns:
        Transformed DataFrame (concatenated if multiple datasets)

    """
    # Load configuration
    if verbose:
        print(f"Loading configuration from: {config_path}")

    dataset_configs, output_path = load_config(config_path)

    if verbose:
        print(f"Found {len(dataset_configs)} dataset configuration(s)")

    # Process each dataset
    all_results = []
    for idx, config in enumerate(dataset_configs, 1):
        if verbose:
            print(f"\n{'='*80}")
            print(f"Processing dataset {idx}/{len(dataset_configs)}")

        # Validate input file exists
        input_path = Path(config.input_file_path)
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_path}")

        # Process dataset
        processor = DatasetProcessor(config)
        result_df = processor.process(verbose=verbose)

        if verbose and config.schema:
            print("\nTransformation Summary:")
            summary = processor.get_summary()
            for key, value in summary.items():
                print(f"  {key}: {value}")

        if len(result_df) > 0:
            all_results.append(result_df)
            if verbose:
                print(f"  Collected {len(result_df)} rows")
        elif verbose:
            print("  No data collected")

    # Concatenate all results
    if all_results:
        final_df = pl.concat(all_results, rechunk=False)
        if verbose:
            print(f"\n{'='*80}")
            print(f"Total rows after concatenation: {len(final_df)}")
    else:
        final_df = pl.DataFrame()
        if verbose:
            print("\nNo data collected from any dataset")

    # Save result if output path is specified
    if output_path:
        output_writer = OutputWriter()
        output_writer.write_csv(final_df, output_path, verbose=verbose)

    return final_df


def main():
    """
    Start the command-line usage from here.

    Returns:
        Exit code (0 for success, 1 for error)

    """
    parser = argparse.ArgumentParser(
        description='Transform datasets into normalized format using YAML configuration.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s config.yaml
  %(prog)s config.yaml --quiet
  %(prog)s config.yaml --show-data
  %(prog)s config.yaml --list-sheets
        """
    )

    parser.add_argument(
        'config',
        type=str,
        help='Path to YAML configuration file'
    )

    parser.add_argument(
        '-q', '--quiet',
        action='store_true',
        help='Suppress progress output'
    )

    parser.add_argument(
        '-s', '--show-data',
        action='store_true',
        help='Display the transformed data'
    )

    parser.add_argument(
        '--head',
        type=int,
        metavar='N',
        help='Show first N rows of output (implies --show-data)'
    )

    parser.add_argument(
        '--list-sheets',
        action='store_true',
        help='List available sheets in Excel file and exit'
    )

    args = parser.parse_args()

    # Handle list-sheets option
    if args.list_sheets:
        try:
            dataset_configs, _ = load_config(args.config)
            config = dataset_configs[0]  # Use first config
            if config.file_type == 'excel':
                sheets = list_excel_sheets(config.input_file_path)
                print(f"\nAvailable sheets in {config.input_file_path}:")
                for i, sheet in enumerate(sheets, 1):
                    print(f"  {i}. {sheet}")
                return 0
            print(f"File is not an Excel file: {config.input_file_path}")
            return 1  # noqa: TRY300
        except Exception as e:
            print(f"Error: {e}", file=sys.stderr)
            import traceback
            traceback.print_exc()
            return 1

    try:
        result_df = process_datasets(args.config, verbose=not args.quiet)

        # Display data if requested
        if args.show_data or args.head:
            print("\n" + "="*80)
            print("Transformed Data:")
            print("="*80)
            if args.head:
                print(result_df.head(args.head))
            else:
                print(result_df)
            return 0
        return 0  # noqa: TRY300

    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
