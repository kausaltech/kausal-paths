# Some functionalities were taken from convert_uploadable_format.py, which was then deleted.
# https://github.com/kausaltech/kausal-paths/blob/91d40f1d3110665f9e80e50ab74fc4d3d7d16b29/notebooks/convert_to_uploadable_format.py

from __future__ import annotations

import argparse
import logging
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, cast

import polars as pl
import yaml
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from nodes.constants import VALUE_COLUMN
from notebooks.notebook_support import get_context

# Import operations from upload_new_dataset.py
from notebooks.upload_new_dataset import (
    canonicalize_metric_column_values,
    clean_dataframe,
    convert_names_to_cats,
    convert_to_standard_format,
    extract_description,
    extract_metrics,
    extract_units,
    extract_units_from_row,
    pivot_by_compound_id,
    prepare_for_dvc,
    process_datasets,
    push_to_dvc,
    to_snake_case,
    write_dataframe_to_csv,
)


@dataclass
class MetricSpec:
    """Defines a metric with its name, ID, unit, column name, and quantity."""

    name: str  # Display name of the metric
    id: str  # Unique identifier (snake_case)
    unit: str  # Unit of measurement
    quantity: str  # Quantity type (e.g., "energy", "emissions")
    column: str | None = None  # Column name in final dataset (defaults to id if not specified)


@dataclass
class ColumnSpec:
    """Describes what a column contains."""

    # For data columns:
    quantity: str = ''
    unit: str = ''
    year: int | None = None

    # For dimension columns (if this column contains dimension values):
    dimension_name: str = ''  # If set, this column contains values for this dimension

    # Additional dimension values (for wide format):
    dimensions: dict[str, str] = field(default_factory=dict)  # dimension_name -> value

    def is_dimension_column(self) -> bool:
        """Check if this column contains dimension values (long format)."""
        return bool(self.dimension_name)

    def is_data_column(self) -> bool:
        """Check if this column contains data values (wide format)."""
        return bool(self.quantity) or bool(self.dimensions)


@dataclass
class DatasetConfig:
    """Complete dataset configuration including file paths and operations."""

    input_file_path: str | None = None  # Required when source='file'; not used for source='dvc' or 'db'
    output_file_path: str | None = None  # Optional, can be set at top level
    column_specs: dict[int, ColumnSpec] | None = None  # column index -> dimensions, quantity, unit (for melt_with_column_specs)
    source: str = 'file'  # 'file' (default), 'dvc', or 'db'
    dataset_id: str | None = None  # Required when source='dvc' or 'db' (e.g. instance_id/dataset_name)
    file_type: str | None = None  # Auto-detected if None: 'csv', 'excel', or 'parquet'
    sheet_name: str | None = None  # For Excel files: specific sheet name
    sheet_year_mapping: dict[str, int] | None = None  # Map sheet names to years
    excel_range: str | None = None  # Excel range like "A3:D20"
    has_header: bool = True  # For file load
    skip_rows: int = 0  # For file load
    operations: list[dict[str, Any]] = field(default_factory=list)  # Sequence of operations to apply
    instance: str | None = None  # Instance ID; required when source='dvc' or 'db'
    metrics: list[MetricSpec] = field(default_factory=list)  # Metric definitions (name, id, unit)
    dataset_name: str | None = None  # Dataset name for DVC metadata (e.g. for push_to_dvc)
    description: str | None = None  # Dataset description for DVC metadata


def parse_excel_range(range_str: str) -> tuple[int, int, int, int]:
    """
    Parse Excel range like "A3:D20" into (start_row, start_col, end_row, end_col).

    Returns 0-based indices: (start_row, start_col, end_row, end_col)
    """

    # Match patterns like "A3:D20" or "A3" (single cell)
    pattern = r'([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?'
    match = re.match(pattern, range_str.upper())
    if not match:
        raise ValueError(f'Invalid Excel range format: {range_str}')

    start_col_str, start_row_str, end_col_str, end_row_str = match.groups()

    # Convert to 0-based indices
    start_col = column_index_from_string(start_col_str) - 1
    start_row = int(start_row_str) - 1

    if end_col_str and end_row_str:
        end_col = column_index_from_string(end_col_str) - 1
        end_row = int(end_row_str) - 1
    else:
        # Single cell - return same as start
        end_col = start_col
        end_row = start_row

    return start_row, start_col, end_row, end_col


class FileLoader:
    """Handles opening and reading files in different formats."""

    @staticmethod
    def _make_column_names_unique(names: list[str]) -> list[str]:
        """Make duplicate column names unique by appending _2, _3, etc. so Polars accepts the schema."""
        seen: dict[str, int] = {}
        result: list[str] = []
        for name in names:
            if name not in seen:
                seen[name] = 1
                result.append(name)
            else:
                seen[name] += 1
                result.append(f'{name}_{seen[name]}')
        return result

    @staticmethod
    def detect_file_type(file_path: str | Path) -> str:
        """Detect file type from extension."""
        path = Path(file_path)
        suffix = path.suffix.lower()
        if suffix in ['.xlsx', '.xls']:
            return 'excel'
        if suffix == '.parquet':
            return 'parquet'
        return 'csv'

    @staticmethod
    def load_csv(file_path: str | Path, skip_rows: int = 0, has_header: bool = True) -> pl.DataFrame:
        """Load a CSV file into a Polars DataFrame."""
        return pl.read_csv(
            file_path,
            skip_rows=skip_rows,
            has_header=has_header,
        )

    @staticmethod
    def load_excel(
        file_path: str | Path,
        sheet_name: str | None = None,
        skip_rows: int = 0,
        has_header: bool = False,
        excel_range: str | None = None,
    ) -> pl.DataFrame:
        """Load an Excel file sheet into a Polars DataFrame."""
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        if ws is None:
            wb.close()
            return pl.DataFrame()

        # Handle Excel range if specified
        if excel_range:
            start_row, start_col, end_row, end_col = parse_excel_range(excel_range)
            # Extract only the specified range
            data = []
            for row_idx in range(start_row, end_row + 1):
                row_data = []
                for col_idx in range(start_col, end_col + 1):
                    cell = ws.cell(row=row_idx + 1, column=col_idx + 1)  # openpyxl uses 1-based
                    row_data.append(cell.value)
                data.append(row_data)
        else:
            # Convert to list of lists
            data = [list(row) for row in ws.iter_rows(values_only=True)]

        wb.close()

        if not data:
            return pl.DataFrame()

        # Determine max columns
        max_cols = max(len(row) for row in data) if data else 0

        # Pad shorter rows
        padded_data = [list(row) + [None] * (max_cols - len(row)) for row in data]

        # Apply skip_rows first
        if skip_rows > 0:
            padded_data = padded_data[skip_rows:]

        if not padded_data:
            return pl.DataFrame()

        # Create column names
        if has_header and len(padded_data) > 0:
            # Use first row as header
            header_row = padded_data[0]
            col_names = [
                f'column_{i + 1}' if (not h or str(h).strip() == '') else str(h).strip()
                for i, h in enumerate(header_row[:max_cols])
            ]
            data_rows = padded_data[1:]
        else:
            col_names = [f'column_{i + 1}' for i in range(max_cols)]
            data_rows = padded_data

        # Make duplicate column names unique so Polars accepts the schema (user can rename later)
        col_names = FileLoader._make_column_names_unique(col_names[:max_cols])

        if not data_rows:
            return pl.DataFrame()

        # Create DataFrame
        df = pl.DataFrame(data_rows, schema=col_names, orient='row')
        return df

    @classmethod
    def load_file(
        cls,
        file_path: str | Path,
        file_type: str | None = None,
        sheet_name: str | None = None,
        skip_rows: int = 0,
        has_header: bool = False,
        excel_range: str | None = None,
    ) -> pl.DataFrame:
        """Load a file based on its type."""
        if file_type is None:
            file_type = cls.detect_file_type(file_path)

        if file_type == 'parquet':
            return cls.load_parquet(file_path)
        if file_type == 'excel':
            return cls.load_excel(file_path, sheet_name, skip_rows, has_header, excel_range)
        # Default to CSV
        return cls.load_csv(file_path, skip_rows, has_header)

    @staticmethod
    def load_parquet(file_path: str | Path) -> pl.DataFrame:
        """Load a parquet file into a Polars DataFrame."""
        return pl.read_parquet(file_path)


def load_from_dvc(instance_id: str, dataset_id: str) -> pl.DataFrame:
    """Load a dataset from DVC. Requires instance context (Django or YAML)."""
    from common import polars as ppl
    from notebooks.notebook_support import get_context

    ctx = get_context(instance_id)
    dvc_ds = ctx.load_dvc_dataset(dataset_id)
    return ppl.from_dvc_dataset(dvc_ds)


def load_from_db(instance_id: str, dataset_id: str) -> pl.DataFrame:
    """Load a dataset from the database for the given instance."""
    from kausal_common.datasets.models import Dataset as DBDatasetModel

    from nodes.datasets import DBDataset
    from nodes.models import InstanceConfig

    ic = InstanceConfig.objects.get(identifier=instance_id)
    db_obj = DBDatasetModel.objects.for_instance_config(ic).get(identifier=dataset_id)  # type: ignore[arg-type]
    return DBDataset.deserialize_df(db_obj)


class OperationsExecutor:
    """Executes operations on DataFrames."""

    def __init__(
        self,
        context=None,
        metrics: list[MetricSpec] | None = None,
        column_specs: dict[int, ColumnSpec] | None = None,
        dataset_name: str | None = None,
        description: str | None = None,
        all_results: list[pl.DataFrame] | None = None,
    ):
        """Initialize with optional context, metrics, column_specs (for melt_with_column_specs), and shared all_results list."""
        self.context = context
        self.metrics = metrics or []
        self.column_specs = column_specs or {}
        self._extracted_units: dict[str, str] = {}  # Store extracted units
        self.dataset_name = dataset_name
        self.description = description
        self.all_results: list[pl.DataFrame] = all_results if all_results is not None else []

    @staticmethod
    def _eval_polars_expr(expr_str: str, _df: pl.DataFrame) -> pl.Expr:
        """
        Safely evaluate a Polars expression string.

        The expression should be a Polars expression like:
        - "pl.col('value') > 0"
        - "pl.col('year') == 2020"
        - "pl.col('name').is_not_null()"
        """
        # Create a safe namespace with only polars
        namespace = {'pl': pl}

        # Compile and evaluate the expression
        try:
            code = compile(expr_str, '<string>', 'eval')
            expr = eval(code, {'__builtins__': {}}, namespace)  # noqa: S307
            if not isinstance(expr, pl.Expr):
                raise TypeError(f'Expression must evaluate to a Polars Expr, got {type(expr)}')  # noqa: TRY301
            return expr  # noqa: TRY300
        except Exception as e:
            raise ValueError(f"Invalid Polars expression '{expr_str}': {e}") from e

    def execute_operation(self, df: pl.DataFrame, operation: dict[str, Any]) -> pl.DataFrame:
        """Execute a single operation on a DataFrame."""
        op_type = operation.get('type')
        op_params = operation.get('params', {})
        method_name = f'_op_{op_type}'
        if not hasattr(self, method_name):
            raise ValueError(f'Unknown operation type: {op_type}')
        return getattr(self, method_name)(df, op_params)

    def execute_operations(self, df: pl.DataFrame, operations: list[dict[str, Any]]) -> pl.DataFrame:
        """Execute a sequence of operations in order."""
        result_df = df
        for operation in operations:
            result_df = self.execute_operation(result_df, operation)
        return result_df

    def _op_append_to_all_results(self, df: pl.DataFrame, _op_params: dict[str, Any]) -> pl.DataFrame:
        """Append the current DataFrame to the shared all_results list. Returns df unchanged."""
        self.all_results.append(df)
        return df

    def _op_clean_dataframe(self, df: pl.DataFrame, _op_params: dict[str, Any]) -> pl.DataFrame:
        return clean_dataframe(df)

    def _op_concat_all_results(self, _df: pl.DataFrame, _op_params: dict[str, Any]) -> pl.DataFrame:
        """
        Replace the current DataFrame with the concatenation of all DataFrames in all_results.

        Column names must match (see Counter check). Categorical columns are cast to Utf8 so they
        align with string columns across frames.
        """
        if not self.all_results:
            return pl.DataFrame()
        frames = self.all_results
        for i, f in enumerate(frames):
            if f.width == 0:
                raise ValueError(f'concat_all_results: frame {i} has no columns.')
        ref_cols = frames[0].columns
        ref_counter = Counter(ref_cols)
        aligned: list[pl.DataFrame] = []
        for i, f in enumerate(frames):
            if Counter(f.columns) != ref_counter:
                msg = (
                    f'concat_all_results: frame {i} columns differ from frame 0 '
                    + '(names and duplicate counts must match; order may differ). '
                    + f'Reference: {ref_cols!r}. This frame: {f.columns!r}.'
                )
                raise ValueError(msg)
            sub = f.select(ref_cols)
            exprs: list[pl.Expr] = []
            for j in range(sub.width):
                if isinstance(sub.to_series(j).dtype, pl.Categorical):
                    exprs.append(pl.nth(j).cast(pl.Utf8))
                else:
                    exprs.append(pl.nth(j))
            aligned.append(sub.select(exprs))
        return pl.concat(aligned, rechunk=False)

    def _op_convert_names_to_cats(self, df: pl.DataFrame, _op_params: dict[str, Any]) -> pl.DataFrame:
        if self.context is None:
            raise ValueError(
                'convert_names_to_cats operation requires a context. ' + "Set 'instance' in the dataset configuration."
            )
        return convert_names_to_cats(df, self.context)

    def _op_convert_to_standard_format(self, df: pl.DataFrame, _op_params: dict[str, Any]) -> pl.DataFrame:
        return convert_to_standard_format(df)

    def _op_define_metrics(self, df: pl.DataFrame, _op_params: dict[str, Any]) -> pl.DataFrame:
        if not self.metrics:
            raise ValueError('define_metrics operation requires metrics to be defined in the dataset configuration.')
        metric_mapping = {}
        for metric in self.metrics:
            column_name = metric.column or metric.id
            metric_mapping[metric.name.lower()] = column_name
            metric_mapping[metric.id.lower()] = column_name

        if 'Metric' in df.columns:
            metric_col_name = 'Metric'
        elif 'Quantity' in df.columns:
            metric_col_name = 'Quantity'
        else:
            raise ValueError("define_metrics operation requires 'Metric' or 'Quantity' column in DataFrame.")

        def map_to_metric_column(name: str | None) -> str:
            if name is None:
                return VALUE_COLUMN
            name_lower = str(name).lower().strip()
            return metric_mapping.get(name_lower, to_snake_case(name))

        return df.with_columns(
            pl.col(metric_col_name).map_elements(map_to_metric_column, return_dtype=pl.Utf8).alias('metric_col')
        )

    def _op_drop(self, df: pl.DataFrame, op_params: dict[str, Any]) -> pl.DataFrame:
        columns = op_params.get('columns', [])
        return df.drop(columns)

    def _op_extract_dimensions_from_text(self, df: pl.DataFrame, op_params: dict[str, Any]) -> pl.DataFrame:  # noqa: C901
        description_col = op_params.get('column', 'Description')
        if description_col not in df.columns:
            raise ValueError(f"extract_dimensions_from_text operation requires '{description_col}' column in DataFrame.")

        category_mapping = op_params.get('mapping')
        if not isinstance(category_mapping, dict):
            raise TypeError(
                "extract_dimensions_from_text operation requires 'mapping' parameter "
                + 'as a dict mapping dimension names to keyword->category_id dictionaries.'
            )

        def find_category_matches(description: str | None, mapping: dict[str, dict[str, str]]) -> dict[str, str | None]:
            matches: dict[str, str | None] = dict.fromkeys(mapping, None)
            if description is None:
                return matches
            description_lower = str(description).lower()
            for dimension_id, keyword_mapping in mapping.items():
                for keyword, category_id in keyword_mapping.items():
                    if keyword.lower() in description_lower:
                        matches[dimension_id] = category_id
                        break
            return matches

        all_dimensions = set(category_mapping.keys())
        if verbose := op_params.get('verbose', False):
            print(f"Extracting dimensions from '{description_col}' column: {sorted(all_dimensions)}")

        df = df.with_columns(
            pl
            .col(description_col)
            .map_elements(
                lambda x: find_category_matches(x, category_mapping),
                return_dtype=pl.Struct([pl.Field(dim, pl.Utf8) for dim in all_dimensions]),
            )
            .alias('_dimension_matches')
        )

        for dimension in all_dimensions:
            df = df.with_columns(pl.col('_dimension_matches').struct.field(dimension).alias(dimension))

        df = df.drop('_dimension_matches')

        if verbose:
            for dimension in sorted(all_dimensions):
                non_null_count = df.filter(pl.col(dimension).is_not_null()).height
                if non_null_count > 0:
                    unique_values = df.select(dimension).unique().drop_nulls()
                    print(f'  {dimension}: {non_null_count} matches, {len(unique_values)} unique values')

        return df

    def _op_extract_metadata(self, df: pl.DataFrame, _op_params: dict[str, Any]) -> pl.DataFrame:
        metadata_cols = ['Quantity', 'Unit', 'Metric', 'metric_col']
        cols_to_drop = [col for col in metadata_cols if col in df.columns]
        if cols_to_drop:
            df = df.drop(cols_to_drop)
        return df

    def _op_extract_units(self, df: pl.DataFrame, _op_params: dict[str, Any]) -> pl.DataFrame:
        if 'metric_col' not in df.columns:
            raise ValueError("extract_units operation requires 'metric_col' column. " + "Run 'define_metrics' operation first.")
        if 'Unit' not in df.columns:
            raise ValueError("extract_units operation requires 'Unit' column in DataFrame.")
        df = canonicalize_metric_column_values(df)
        units = extract_units(df)
        self._extracted_units = units
        return df

    def _op_extract_units_from_row(self, df: pl.DataFrame, _op_params: dict[str, Any]) -> pl.DataFrame:
        units, df_cleaned = extract_units_from_row(df)
        self._extracted_units = units
        return df_cleaned

    def _op_filter(self, df: pl.DataFrame, op_params: dict[str, Any]) -> pl.DataFrame:
        expr_str = op_params.get('expr')
        if not expr_str:
            raise ValueError("Filter operation requires 'expr' parameter")
        expr = self._eval_polars_expr(expr_str, df)
        return df.filter(expr)

    def _op_filter_by_sector_carrier(self, df: pl.DataFrame, op_params: dict[str, Any]) -> pl.DataFrame:
        """
        Filter rows to those whose column (e.g. field_key) is in the sector * energy_carrier set.

        Params:
            column: Column name (default "field_key").
            sectors: List of sector prefixes (e.g. ["EW_", "GE_", "HE_", "IE_", "KE_"]).
            energy_carriers: List of energy carrier names (e.g. ["Biogas", "Biomasse", ...]).
            include: If True (default), keep only matching rows; if False, drop matching rows.
        """
        column = op_params.get('column', 'field_key')
        sectors = op_params.get('sectors')
        energy_carriers = op_params.get('energy_carriers')
        if not sectors or not energy_carriers:
            raise ValueError("filter_by_sector_carrier requires 'sectors' and 'energy_carriers' lists in params.")
        if column not in df.columns:
            raise ValueError(f'filter_by_sector_carrier: column {column!r} not in DataFrame (columns: {list(df.columns)})')
        allowed = build_sector_carrier_field_keys(sectors, energy_carriers)
        include = op_params.get('include', True)
        if include:
            return df.filter(pl.col(column).is_in(allowed))
        return df.filter(~pl.col(column).is_in(allowed))

    def _op_melt_with_column_specs(self, df: pl.DataFrame, op_params: dict[str, Any]) -> pl.DataFrame:
        """
        Melt value columns to long format and add dimension/quantity/unit columns from column_specs.

        Uses dataset-level column_specs (or params override). For each melted column, looks up
        dimensions, quantity, and unit from the spec and adds them as columns. Quantity/unit
        fall back to the first dataset-level metric if not set on the spec.

        Params:
            id_vars: Column names to keep as identifiers (default: all columns not in column_specs).
            value_name: Name for the value column (default 'Value').
            variable_column_name: Name for the column that identifies the source column (default '_column').
            drop_variable_column: If True (default), drop the variable column after joining specs.
        """
        specs = op_params.get('column_specs')
        if specs is not None:
            # Params can pass column_specs as dict of int -> dict (same shape as YAML)
            specs = parse_column_specs({'column_specs': specs})
        else:
            specs = self.column_specs
        if not specs:
            raise ValueError("melt_with_column_specs requires 'column_specs' in dataset config or in params.")
        default_quantity = self.metrics[0].quantity if self.metrics else ''
        default_unit = self.metrics[0].unit if self.metrics else ''
        value_name = op_params.get('value_name', 'Value')
        variable_column_name = op_params.get('variable_column_name', '_column')
        drop_variable = op_params.get('drop_variable_column', True)

        # Data columns only (specs that have dimensions or quantity)
        value_indices = sorted(i for i in specs if i < len(df.columns) and specs[i].is_data_column())
        if not value_indices:
            raise ValueError('melt_with_column_specs: no column_specs entries are data columns (need dimensions or quantity).')
        value_vars = [df.columns[i] for i in value_indices]
        id_vars = op_params.get('id_vars')
        if id_vars is None:
            id_vars = [c for j, c in enumerate(df.columns) if j not in specs]
        if not id_vars:
            raise ValueError(
                'melt_with_column_specs: need at least one id column (set id_vars or use column_specs for subset of columns).'
            )

        long_df = df.unpivot(
            index=id_vars,
            on=value_vars,
            variable_name=variable_column_name,
            value_name=value_name,
        )

        # Build lookup: variable column value -> Quantity, Unit, and dimension columns
        lookup_rows = []
        for idx in value_indices:
            spec = specs[idx]
            quantity = spec.quantity or default_quantity
            unit = spec.unit or default_unit
            row: dict[str, Any] = {
                variable_column_name: df.columns[idx],
                'Quantity': quantity,
                'Unit': unit,
                **spec.dimensions,
            }
            if spec.year is not None:
                row['Year'] = spec.year
            lookup_rows.append(row)
        lookup_df = pl.DataFrame(lookup_rows)

        result = long_df.join(lookup_df, on=variable_column_name, how='left')
        if drop_variable:
            result = result.drop(variable_column_name)
        return result

    def _op_pivot_by_compound_id(self, df: pl.DataFrame, _op_params: dict[str, Any]) -> pl.DataFrame:
        return pivot_by_compound_id(df)

    def _op_prepare_for_dvc(self, df: pl.DataFrame, op_params: dict[str, Any]) -> pl.DataFrame:
        units = op_params.get('units', {})
        return prepare_for_dvc(df, units)

    def _op_print_data(self, df: pl.DataFrame, _op_params: dict[str, Any]) -> pl.DataFrame:
        print(df)
        return df

    def _op_print_metadata(self, df: pl.DataFrame, op_params: dict[str, Any]) -> pl.DataFrame:  # noqa: C901, PLR0912, PLR0915
        print('\n' + '=' * 80)
        print('METADATA SUMMARY')
        print('=' * 80)

        units_to_print = op_params.get('units')
        if units_to_print is None:
            units_to_print = self._extracted_units
        if units_to_print:
            print(f'\nUnits ({len(units_to_print)}):')
            print('-' * 80)
            for metric_id, unit in sorted(units_to_print.items()):
                print(f'  {metric_id:30s} -> {unit}')
        else:
            print('\nUnits: None extracted')

        if self.metrics:
            print(f'\nMetrics ({len(self.metrics)}):')
            print('-' * 80)
            for metric in self.metrics:
                column_name = metric.column or metric.id
                print(f'  Name:     {metric.name}')
                print(f'    ID:     {metric.id}')
                print(f'    Column: {column_name}')
                print(f'    Unit:   {metric.unit}')
                print(f'    Quantity: {metric.quantity}')
                print()
        else:
            print('\nMetrics: None defined in configuration')

        print('DataFrame Structure:')
        print('-' * 80)
        print(f'  Rows: {len(df)}')
        print(f'  Columns ({len(df.columns)}):')
        metric_cols = []
        dimension_cols = []
        other_cols = []
        units_dict = units_to_print or {}
        # Columns that are metrics (from YAML metric definitions) are metric columns
        metric_column_names = {m.column or m.id for m in self.metrics}
        for col in df.columns:
            if col in units_dict or col in metric_column_names:
                metric_cols.append(col)
            elif col.lower() == 'year':
                other_cols.append(f'{col} (time dimension)')
            else:
                dimension_cols.append(col)
        if metric_cols:
            print(f'    Metric columns ({len(metric_cols)}): {", ".join(sorted(metric_cols))}')
        if dimension_cols:
            print(f'    Dimension columns ({len(dimension_cols)}): {", ".join(sorted(dimension_cols))}')
        if other_cols:
            print(f'    Other columns: {", ".join(other_cols)}')
        print('  Column dtypes:')
        for col in df.columns:
            dtype = df.schema[col]
            print(f'    {col:30s} -> {dtype}')

        if len(df) > 0:
            sample_size = op_params.get('sample_rows', 5)
            print(f'\nSample Data (first {min(sample_size, len(df))} rows):')
            print('-' * 80)
            sample_df = df.head(sample_size)
            for idx, row in enumerate(sample_df.iter_rows(named=True), 1):
                print(f'  Row {idx}:')
                for key, value in row.items():
                    if value is not None:
                        print(f'    {key}: {value}')
                print()

        print('=' * 80 + '\n')
        print(df.describe())
        return df

    def _op_process_datasets(self, df: pl.DataFrame, op_params: dict[str, Any]) -> pl.DataFrame:
        outcsvpath = op_params.get('outcsvpath', '')
        output_path = op_params.get('output_path')
        if not output_path:
            raise ValueError("process_datasets operation requires 'output_path' parameter")
        dataset_name = op_params.get('dataset_name') or self.dataset_name or None
        language = op_params.get('language', 'en')
        process_datasets(df, outcsvpath, output_path, language, self.context, dataset_name)
        return df

    def _op_push_to_dvc(self, df: pl.DataFrame, op_params: dict[str, Any]) -> pl.DataFrame:  # noqa: C901, PLR0912
        from nodes.node import NodeMetric

        output_path = op_params.get('output_path')
        if not output_path:
            raise ValueError("push_to_dvc operation requires 'output_path' parameter")

        dataset_name = op_params.get('dataset_name') or self.dataset_name or 'dataset'
        language = op_params.get('language', 'en')

        # Prefer units from PathsDataFrame meta when available
        units = op_params.get('units')
        if units is None:
            get_meta = getattr(df, 'get_meta', None)
            if callable(get_meta):
                meta = get_meta()
                meta_units = getattr(meta, 'units', {})
                units = {col: str(u) for col, u in meta_units.items()}
        if units is None:
            if self.metrics:
                units = {}
                for metric_spec in self.metrics:
                    column_name = metric_spec.column or metric_spec.id
                    units[column_name] = metric_spec.unit
            else:
                units = self._extracted_units
        if not isinstance(units, dict):
            raise TypeError("push_to_dvc operation requires 'units' to be a dict")

        if not units:
            print('Warning: No units found. Units should be:')
            print("  1. Provided via params['units']")
            print('  2. From PathsDataFrame meta (e.g. after to_paths_dataframe or load from DVC/DB)')
            print('  3. Defined in YAML metrics (metric.unit)')
            print("  4. Extracted via 'extract_units' operation (before pivot/extract_metadata)")

        description = op_params.get('description') or self.description
        if description is None and 'Description' in df.columns:
            description = extract_description(df)

        node_metrics: list[NodeMetric] = []
        if self.metrics:
            for metric_spec in self.metrics:
                column_id = metric_spec.column or metric_spec.id
                node_metrics.append(
                    NodeMetric(
                        unit=metric_spec.unit,
                        quantity=to_snake_case(metric_spec.quantity),
                        id=metric_spec.id,
                        label=cast('Any', {language: metric_spec.name}),
                        column_id=column_id,
                    )
                )
        elif op_params.get('extract_metrics', False):
            if 'metric_col' in df.columns and 'Metric' in df.columns and 'Quantity' in df.columns:
                df = canonicalize_metric_column_values(df)
                extracted_units = extract_units(df)
                node_metrics = extract_metrics(df, language, extracted_units)
                units = extracted_units
            else:
                print('Warning: Cannot extract metrics - missing required columns (metric_col, Metric, Quantity)')

        push_to_dvc(
            df=df,
            output_path=output_path,
            dataset_name=dataset_name,
            description=description,
            metrics=node_metrics,
            language=language,
            units=units,
        )
        print(f"✓ Dataset '{dataset_name}' pushed to DVC at {output_path}")
        return df

    def _op_rename(self, df: pl.DataFrame, op_params: dict[str, Any]) -> pl.DataFrame:
        mapping = op_params.get('mapping', {})
        return df.rename(mapping)

    def _op_select(self, df: pl.DataFrame, op_params: dict[str, Any]) -> pl.DataFrame:
        columns = op_params.get('columns', [])
        return df.select(columns)

    def _op_set_unit(self, df: pl.DataFrame, op_params: dict[str, Any]) -> pl.DataFrame:
        """Set or change units on metric columns. Only applies when df is a PathsDataFrame (e.g. from DVC/DB)."""
        set_unit_fn = getattr(df, 'set_unit', None)
        if not callable(set_unit_fn):
            raise TypeError(
                'set_unit operation requires a PathsDataFrame (e.g. load with source: dvc or source: db). '
                + 'File-loaded data does not carry unit metadata; use metrics in YAML or push_to_dvc params instead.'
            )
        mapping = op_params.get('mapping', {})
        if not mapping:
            col = op_params.get('column')
            unit = op_params.get('unit')
            if col is None or unit is None:
                raise ValueError("set_unit requires params 'mapping' (dict) or 'column' and 'unit'.")
            mapping = {col: unit}
        for col, unit_str in mapping.items():
            if col not in df.columns:
                raise ValueError(f'set_unit: column {col!r} not in DataFrame.')
            df = cast('pl.DataFrame', set_unit_fn(col, unit_str, force=True))
        return df

    def _op_sort(self, df: pl.DataFrame, op_params: dict[str, Any]) -> pl.DataFrame:
        """Sort rows by column(s). Params: by (column name or list), descending (bool or list), nulls_last (bool)."""
        by = op_params.get('by')
        if by is None:
            raise ValueError("sort operation requires 'by' parameter (column name or list of column names)")
        if isinstance(by, str):
            by = [by]
        descending = op_params.get('descending', False)
        nulls_last = op_params.get('nulls_last', True)
        return df.sort(by, descending=descending, nulls_last=nulls_last)

    def _op_to_paths_dataframe(self, df: pl.DataFrame, op_params: dict[str, Any]) -> pl.DataFrame:
        """Convert pl.DataFrame to PathsDataFrame with given units and primary keys (step 4 in pipeline)."""
        from common import polars as ppl
        from nodes.units import unit_registry

        units_dict = op_params.get('units', {})
        primary_keys = op_params.get('primary_keys', [])
        if not isinstance(units_dict, dict):
            raise TypeError("to_paths_dataframe requires 'units' to be a dict")
        units_parsed = {col: unit_registry.parse_units(u) for col, u in units_dict.items()}
        meta = ppl.DataFrameMeta(units=units_parsed, primary_keys=list(primary_keys))
        return ppl.to_ppdf(df, meta=meta)

    def _op_to_snake_case_columns(self, df: pl.DataFrame, op_params: dict[str, Any]) -> pl.DataFrame:
        exclude = op_params.get('exclude', ['Year'])
        new_cols = {}
        for col in df.columns:
            if col not in exclude:
                new_cols[col] = to_snake_case(col)
        return df.rename(new_cols)

    def _op_with_columns(self, df: pl.DataFrame, op_params: dict[str, Any]) -> pl.DataFrame:
        expr_or_exprs = op_params.get('expr') or op_params.get('exprs')
        if expr_or_exprs is None:
            raise ValueError("with_columns operation requires 'expr' or 'exprs' parameter")
        if isinstance(expr_or_exprs, str):
            expr = self._eval_polars_expr(expr_or_exprs, df)
            return df.with_columns(expr)
        if isinstance(expr_or_exprs, list):
            exprs = [self._eval_polars_expr(e, df) for e in expr_or_exprs]
            return df.with_columns(exprs)
        raise ValueError("with_columns 'expr'/'exprs' must be a string or list of strings")

    def _op_write_csv(self, df: pl.DataFrame, op_params: dict[str, Any]) -> pl.DataFrame:
        """Write the current DataFrame to a CSV file. Returns the DataFrame unchanged."""
        output_path = op_params.get('output_path')
        if not output_path:
            raise ValueError("write_csv operation requires 'output_path' parameter")
        verbose = op_params.get('verbose', True)
        write_dataframe_to_csv(df, output_path, verbose=verbose)
        return df


class DatasetProcessor:
    """Orchestrates the complete dataset processing pipeline."""

    def __init__(self, config: DatasetConfig, all_results: list[pl.DataFrame] | None = None):
        self.config = config
        self.file_loader = FileLoader()
        self._all_results: list[pl.DataFrame] = all_results if all_results is not None else []
        context = None
        if config.instance:
            try:
                context = get_context(config.instance)
            except Exception as e:
                print(f"Warning: Could not load context for instance '{config.instance}': {e}")
        self.operations_executor = OperationsExecutor(
            context=context,
            metrics=config.metrics,
            column_specs=config.column_specs or {},
            dataset_name=config.dataset_name,
            description=config.description,
            all_results=self._all_results,
        )

    def process(self, verbose: bool = True) -> pl.DataFrame:
        """Process a single dataset configuration: load from file/DVC/DB, then apply operations."""
        if self.config.source == 'file':
            assert self.config.input_file_path is not None
            if verbose:
                print(f'Loading file: {self.config.input_file_path}')
            df = self.file_loader.load_file(
                self.config.input_file_path,
                self.config.file_type,
                sheet_name=self.config.sheet_name,
                skip_rows=self.config.skip_rows,
                has_header=self.config.has_header,
                excel_range=self.config.excel_range,
            )
        elif self.config.source == 'dvc':
            assert self.config.instance is not None
            assert self.config.dataset_id is not None
            if verbose:
                print(f'Loading from DVC: {self.config.dataset_id} (instance: {self.config.instance})')
            df = load_from_dvc(self.config.instance, self.config.dataset_id)
        elif self.config.source == 'db':
            assert self.config.instance is not None
            assert self.config.dataset_id is not None
            if verbose:
                print(f'Loading from database: {self.config.dataset_id} (instance: {self.config.instance})')
            df = load_from_db(self.config.instance, self.config.dataset_id)
        else:
            raise ValueError(f'Unknown source: {self.config.source!r}')

        get_meta = getattr(df, 'get_meta', None)
        saved_meta = get_meta() if callable(get_meta) else None

        if self.config.operations:
            df = self.operations_executor.execute_operations(df, self.config.operations)

        if saved_meta is not None:
            from common import polars as ppl

            df = ppl.to_ppdf(df, meta=cast('ppl.DataFrameMeta | None', saved_meta))

        return df

    def get_summary(self) -> dict[str, Any]:
        """Return a short summary when column_specs are defined."""
        if not self.config.column_specs:
            return {}
        data_cols = [s for s in self.config.column_specs.values() if s.is_data_column()]
        return {
            'column_specs': len(self.config.column_specs),
            'data_columns': len(data_cols),
        }


def build_sector_carrier_field_keys(
    sectors: list[str],
    energy_carriers: list[str],
) -> list[str]:
    """
    Build the full list of field_key values for sector * energy_carrier combinations.

    Each value is sector_prefix + energy_carrier (e.g. EW_Biogas, KE_Umweltwaerme).
    Use with filter operation: pl.col('field_key').is_in(build_sector_carrier_field_keys(...))
    or with the filter_by_sector_carrier operation in YAML.
    """
    return [s + c for s in sectors for c in energy_carriers]


def list_excel_sheets(file_path: str | Path) -> list[str]:
    """List all sheet names in an Excel file."""
    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f'File not found: {file_path}')

    if file_path.suffix.lower() not in ['.xlsx', '.xls']:
        raise ValueError(f'File is not an Excel file: {file_path}')

    wb = load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    return sheet_names


def merge_configs(base_config: dict[str, Any], override_config: dict[str, Any]) -> dict[str, Any]:
    """Merge two config dictionaries, with override taking precedence."""
    merged = base_config.copy()

    # Deep merge for nested dictionaries
    for key, value in override_config.items():
        if key in merged and isinstance(merged[key], dict) and isinstance(value, dict):
            merged[key] = merge_configs(merged[key], value)
        elif key in merged and isinstance(merged[key], list) and isinstance(value, list):
            # For lists, replace entirely (don't merge)
            merged[key] = value
        else:
            merged[key] = value

    return merged


def parse_column_specs(config_dict: dict[str, Any]) -> dict[int, ColumnSpec]:
    """Parse column specifications from config dictionary."""
    column_specs = {}
    for col_num_str, spec_dict in config_dict.get('column_specs', {}).items():
        col_num = int(col_num_str)

        # Create ColumnSpec based on whether it's a dimension or data column
        if spec_dict.get('dimension_name'):
            # This is a dimension column (long format)
            column_specs[col_num] = ColumnSpec(dimension_name=spec_dict['dimension_name'])
        else:
            # This is a data column (wide format)
            column_specs[col_num] = ColumnSpec(
                quantity=spec_dict.get('quantity', ''),
                unit=spec_dict.get('unit', ''),
                year=spec_dict.get('year'),
                dimensions=spec_dict.get('dimensions', {}),
            )
    return column_specs


def parse_metrics(config_list: list[dict[str, Any]] | None) -> list[MetricSpec]:
    """Parse metric specifications from config list."""
    if not config_list:
        return []
    return [
        MetricSpec(
            name=metric_dict['name'],
            id=metric_dict['id'],
            unit=metric_dict['unit'],
            quantity=metric_dict['quantity'],
            column=metric_dict.get('column'),  # Optional, defaults to None
        )
        for metric_dict in config_list
    ]


def parse_dataset_config(
    config_dict: dict[str, Any], default_output: str | None = None, base_config: dict[str, Any] | None = None
) -> DatasetConfig:
    """Parse a single dataset configuration from dictionary."""
    # Merge with base config if provided (property inheritance)
    if base_config:
        config_dict = merge_configs(base_config, config_dict)

    source = config_dict.get('source', 'file')
    if source not in ('file', 'dvc', 'db'):
        raise ValueError(f"Invalid source: {source!r}. Must be 'file', 'dvc', or 'db'.")

    input_file_path = config_dict.get('input_file_path')
    dataset_id = config_dict.get('dataset_id')

    if source == 'file':
        if not input_file_path:
            raise ValueError("input_file_path is required when source is 'file'.")
        file_type = config_dict.get('file_type') or FileLoader.detect_file_type(Path(input_file_path))
    else:
        if not dataset_id:
            raise ValueError(f'dataset_id is required when source is {source!r}.')
        if not config_dict.get('instance'):
            raise ValueError(f'instance is required when source is {source!r}.')
        file_type = None

    # Parse column_specs and metrics if present
    column_specs = parse_column_specs(config_dict) if 'column_specs' in config_dict else None
    metrics = parse_metrics(config_dict.get('metrics'))

    return DatasetConfig(
        input_file_path=input_file_path,
        output_file_path=config_dict.get('output_file_path', default_output),
        column_specs=column_specs,
        source=source,
        dataset_id=dataset_id,
        file_type=file_type,
        sheet_name=config_dict.get('sheet_name'),
        sheet_year_mapping=config_dict.get('sheet_year_mapping'),
        excel_range=config_dict.get('excel_range'),
        has_header=config_dict.get('has_header', True),
        skip_rows=config_dict.get('skip_rows', 0),
        operations=config_dict.get('operations', []),
        instance=config_dict.get('instance'),
        metrics=metrics,
        dataset_name=config_dict.get('dataset_name'),
        description=config_dict.get('description'),
    )


def load_config(yaml_file_path: str | Path) -> tuple[list[DatasetConfig], str | None]:
    """
    Load dataset configurations from YAML file.

    Supports both single config and list of configs. No operations are injected;
    use append_to_all_results, concat_all_results, and write_csv explicitly in your YAML.

    Args:
        yaml_file_path: Path to the YAML configuration file

    Returns:
        Tuple of (list of DatasetConfig objects, output_file_path from config if set).

    """
    yaml_file_path = Path(yaml_file_path)

    if not yaml_file_path.exists():
        raise FileNotFoundError(f'Configuration file not found: {yaml_file_path}')

    with yaml_file_path.open('r', encoding='utf-8') as f:
        config = yaml.safe_load(f)

    # Handle top-level output path, instance, and metrics
    default_output: str | None = config.get('output_file_path')
    top_level_instance: str | None = config.get('instance')
    top_level_metrics: list[dict[str, Any]] | None = config.get('metrics')

    # Check if config is a list of datasets
    if 'datasets' in config:
        # No inheritance between datasets: each is parsed with the same top-level defaults only.
        # TODO: Add inheritance between datasets and/or one dataset from several input files
        # (e.g. input_file_paths: [...]) when there is a concrete use case to design and test.
        dataset_configs = []
        top_level_base: dict[str, Any] = {}
        if top_level_instance is not None:
            top_level_base['instance'] = top_level_instance
        if top_level_metrics is not None:
            top_level_base['metrics'] = top_level_metrics
        if default_output is not None:
            top_level_base['output_file_path'] = default_output

        for dataset_dict in config['datasets']:
            dataset_config = parse_dataset_config(dataset_dict, default_output, top_level_base or None)
            dataset_configs.append(dataset_config)
        return dataset_configs, default_output

    # Single dataset configuration
    single_base_config: dict[str, Any] = {}
    if top_level_instance is not None:
        single_base_config['instance'] = top_level_instance
    if top_level_metrics is not None:
        single_base_config['metrics'] = top_level_metrics
    dataset_config = parse_dataset_config(config, default_output, single_base_config or None)
    return [dataset_config], default_output


def process_yaml_datasets(config_path: str | Path, verbose: bool = True) -> pl.DataFrame:
    """
    Process one or more datasets using configuration from YAML file.

    Args:
        config_path: Path to YAML configuration file
        verbose: Whether to print progress information

    Returns:
        Transformed DataFrame (concatenated if multiple datasets)

    """
    # Load configuration
    if verbose:
        print(f'Loading configuration from: {config_path}')

    dataset_configs, _ = load_config(config_path)

    if verbose:
        print(f'Found {len(dataset_configs)} dataset configuration(s)')

    # Shared list for append_to_all_results; same list is passed to every processor
    all_results: list[pl.DataFrame] = []
    last_result: pl.DataFrame = pl.DataFrame()

    for idx, config in enumerate(dataset_configs, 1):
        if verbose:
            print(f'\n{"=" * 80}')
            print(f'Processing dataset {idx}/{len(dataset_configs)}')

        # Validate input file exists when loading from file
        if config.source == 'file' and config.input_file_path:
            input_path = Path(config.input_file_path)
            if not input_path.exists():
                raise FileNotFoundError(f'Input file not found: {input_path}')

        processor = DatasetProcessor(config, all_results=all_results)
        result_df = processor.process(verbose=verbose)
        last_result = result_df

        if verbose and config.column_specs:
            print('\nColumn specs summary:')
            summary = processor.get_summary()
            for key, value in summary.items():
                print(f'  {key}: {value}')

        if verbose and len(result_df) > 0:
            print(f'  Collected {len(result_df)} rows')

    return last_result


def main():
    """
    Start the command-line usage from here.

    Returns:
        Exit code (0 for success, 1 for error)

    """
    import os

    os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'paths.settings')
    import django

    django.setup()

    parser = argparse.ArgumentParser(
        description='Transform datasets into normalized format using YAML configuration.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s config.yaml
  %(prog)s config.yaml --quiet
  %(prog)s config.yaml --show-data
  %(prog)s config.yaml --list-sheets
        """,
    )

    parser.add_argument('config', type=str, help='Path to YAML configuration file')

    parser.add_argument('-q', '--quiet', action='store_true', help='Suppress progress output')

    parser.add_argument('-s', '--show-data', action='store_true', help='Display the transformed data')

    parser.add_argument('--head', type=int, metavar='N', help='Show first N rows of output (implies --show-data)')

    parser.add_argument('--list-sheets', action='store_true', help='List available sheets in Excel file and exit')

    args = parser.parse_args()

    # Reduce log noise: default INFO (no DEBUG from nodes/dvc_pandas/dulwich), or WARNING if -q
    from loguru import logger

    logger.remove()
    logger.add(sys.stderr, level='WARNING' if args.quiet else 'INFO')
    logging.getLogger().setLevel(logging.WARNING if args.quiet else logging.INFO)

    # Handle list-sheets option
    if args.list_sheets:
        try:
            dataset_configs, _ = load_config(args.config)
            config = dataset_configs[0]  # Use first config
            if config.source != 'file' or not config.input_file_path:
                print("List-sheets is only supported for source='file' with input_file_path.", file=sys.stderr)
                return 1
            if config.file_type == 'excel':
                sheets = list_excel_sheets(config.input_file_path)
                print(f'\nAvailable sheets in {config.input_file_path}:')
                for i, sheet in enumerate(sheets, 1):
                    print(f'  {i}. {sheet}')
                return 0
            print(f'File is not an Excel file: {config.input_file_path}')
            return 1  # noqa: TRY300
        except Exception as e:
            print(f'Error: {e}', file=sys.stderr)
            import traceback

            traceback.print_exc()
            return 1

    try:
        result_df = process_yaml_datasets(args.config, verbose=not args.quiet)

        # Display data if requested
        if args.show_data or args.head:
            print('\n' + '=' * 80)
            print('Transformed Data:')
            print('=' * 80)
            if args.head:
                print(result_df.head(args.head))
            else:
                print(result_df)
            return 0
        return 0  # noqa: TRY300

    except Exception as e:
        print(f'Error: {e}', file=sys.stderr)
        import traceback

        traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(main())
