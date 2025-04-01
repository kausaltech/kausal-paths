from __future__ import annotations

import hashlib
import re
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING, ClassVar

import polars as pl
import requests
from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import absolute_coordinate, get_column_letter, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName

from kausal_common.debugging.perf import PerfCounter

from paths.const import MODEL_CALC_OP

from common.i18n import I18nBaseModel, I18nStringInstance
from nodes.constants import BASELINE_SCENARIO, FORECAST_COLUMN, IMPACT_COLUMN, IMPACT_GROUP, VALUE_COLUMN, YEAR_COLUMN

if TYPE_CHECKING:
    from openpyxl.cell.cell import Cell
    from openpyxl.worksheet.worksheet import Worksheet

    from nodes.actions.action import ActionNode
    from nodes.context import Context
    from nodes.dimensions import Dimension
    from nodes.instance import Instance
    from nodes.models import InstanceConfig
    from nodes.node import Node


DATA_SHEET_NAME = 'Data'
PARAM_SHEET_NAME = 'Parameters'

FIXED_COLUMNS_LONG = [
    'Node',
    'Year',
    'Unit',
    'Value',
    'BaselineValue',
    'Forecast',
]

FIXED_COLUMNS_WIDE = [
    'Node',
    'Quantity',
    'Unit',
    'Forecast_from',
]


class ExportNotSupportedError(Exception):
    pass


def _convert_to_camel_case(input_string: str) -> str:
    # Split the input string into words
    words = re.split(r'[-\s]+', input_string)

    # Capitalize the first letter of each word
    capitalized_words = [word.capitalize() for word in words]

    # Join the words without spaces
    camel_case_string = ''.join(capitalized_words)

    return camel_case_string


def _dim_to_col(dim: Dimension) -> str:
    return _convert_to_camel_case(str(dim.label))


class InstanceResultExcel(I18nBaseModel):
    name: I18nStringInstance
    base_excel_url: str | None
    node_ids: list[str] | None = None
    action_ids: list[str] | None = None
    format: str = 'long'

    _created_sheet_names: ClassVar = (DATA_SHEET_NAME, PARAM_SHEET_NAME)

    def validate_for_instance(self, instance: Instance):
        ctx = instance.context
        if self.node_ids is not None:
            for node_id in self.node_ids:
                if node_id not in ctx.nodes:
                    raise KeyError(f"Node {node_id} not found.")
        if self.action_ids is not None:
            actions = {n.id for n in ctx.get_actions()}
            for action_id in self.action_ids:
                if action_id not in actions:
                    raise KeyError(f"Action {action_id} not found.")

    def _output_node_long(
        self,
        context: Context,
        wb: Workbook,
        sheet: Worksheet,
        node: Node,
        dim_ids: list[str],
        actions: list[ActionNode],
        aseq: bool,
    ) -> None:
        logger.info('Outputting node %s' % node.id)
        if len(node.output_metrics) > 1:
            logger.warning('Multimetric node %s' % node.id)
        df = node.get_output_pl()
        if df.dim_ids:
            df = df.with_columns([pl.col(dim_id).cast(pl.String) for dim_id in df.dim_ids])
        df = df.sort(by=[YEAR_COLUMN, *df.dim_ids])
        bdf = node.get_baseline_values()
        assert bdf is not None
        bdf = bdf.with_columns(pl.col(VALUE_COLUMN).alias('BaselineValue')).drop(VALUE_COLUMN)
        df = df.paths.join_over_index(bdf, how='left', index_from='left')
        col_map = {
            'Node': pl.lit(node.id),
            'Year': pl.col(YEAR_COLUMN),
            'Unit': pl.lit(str(node.unit)),
            'Value': pl.col(VALUE_COLUMN),
            'BaselineValue': pl.col('BaselineValue'),
            'Forecast': pl.col(FORECAST_COLUMN),
        }
        cols = [col_map[col_id].alias(col_id) for col_id in FIXED_COLUMNS_LONG]
        for dim_id in dim_ids:
            if dim_id in df.dim_ids:
                cols.append(pl.col(dim_id))
            else:
                cols.append(pl.lit(None).alias(dim_id))

        arange = range(len(actions))
        for i in arange:
            if aseq:
                for j in arange:
                    actions[j].enabled_param.set(j <= i)

            with context.start_span('compute impact: %s' % actions[i].id, op='function'):
                adf = actions[i].compute_impact(node)
            act_col = 'Impact_%s' % actions[i].id
            adf = adf.filter(pl.col(IMPACT_COLUMN) == IMPACT_GROUP).drop(IMPACT_COLUMN).rename({VALUE_COLUMN: act_col})
            df = df.paths.join_over_index(adf, how='left', index_from='left')
            cols.append(pl.col(act_col))

        start_row = sheet.max_row + 1

        df = df.select(cols)
        for row in df.iter_rows():
            sheet.append(row)
        end_row = sheet.max_row

        start_col = get_column_letter(1)
        end_col = get_column_letter(len(row))  # type: ignore
        cell_range = absolute_coordinate('%s%s:%s%s' % (start_col, start_row, end_col, end_row))
        ref = '%s!%s' % (quote_sheetname(sheet.title), cell_range)
        defn = DefinedName(node.id, attr_text=ref)
        wb.defined_names[node.id] = defn

    def _output_node_wide(
        self,
        sheet: Worksheet,
        node: Node,
        dim_ids: list[str],
        cols: list[str],
    ) -> None:
        logger.info('Outputting node %s' % node.id)
        df = node.get_output_pl()
        if df.dim_ids:
            df = df.with_columns([pl.col(dim_id).cast(pl.String) for dim_id in df.dim_ids])

        # Get baseline values
        bdf = node.get_baseline_values()
        assert bdf is not None

        # Create the basic structure
        col_map = {
            'Node': pl.lit(node.id),
            'Quantity': pl.lit(node.quantity),
            'Unit': pl.lit(str(node.unit)),
            'Forecast_from': pl.col('Forecast_from'),
        }

        # Add dimension columns
        for dim_id in dim_ids:
            if dim_id in df.dim_ids:
                col_map[dim_id] = pl.col(dim_id)
            else:
                col_map[dim_id] = pl.lit(None).alias(dim_id)

        # Add a dummy grouping column as at least one is required
        df = df.with_columns(pl.lit(1).alias('_dummy_group'))
        grouping_cols = df.dim_ids if df.dim_ids else ['_dummy_group']

        # Create forecast_from values
        forecast_from_df = (df
            .filter(pl.col(FORECAST_COLUMN))
            .group_by(grouping_cols)
            .agg(
                pl.col(YEAR_COLUMN).min().alias('Forecast_from')
            )
        )

        dfout: pl.DataFrame = df.join(
            forecast_from_df,
            on=grouping_cols,
            how='left'
        ).drop('_dummy_group')  # Remove the dummy column after join

        dfout = dfout.select([
                *[col_map[col_id].alias(col_id) for col_id in col_map.keys()],
                pl.col(YEAR_COLUMN),
                pl.col(VALUE_COLUMN),
                pl.col(FORECAST_COLUMN)
            ]).pivot(  # type: ignore[call-arg]
                values=VALUE_COLUMN,
                index=[*col_map.keys()],  # Now includes Forecast_from
                columns=YEAR_COLUMN,
                maintain_order=True,
                aggregate_function='first'
            )

        dfout = dfout.with_columns([
            pl.col(col).cast(pl.String, strict=False).fill_null(".")
            for col in dfout.columns
            if col in dim_ids
        ])
        for col in cols:
            if str(col) not in dfout.columns:
                dfout = dfout.with_columns(pl.lit(None).alias(str(col)))
        dfout = dfout.select(cols)

        # Write to sheet
        for row in dfout.iter_rows():
            sheet.append(row)

    def _add_param_sheet(self, context: Context, wb: Workbook) -> None:
        ps: Worksheet = wb.create_sheet(PARAM_SHEET_NAME)
        ps.append(['Parameter', 'Value'])

        max_name_length: int = 20

        def add_param_value(name: str, value: float | str | None, named_range_name: str | None = None) -> None:
            nonlocal max_name_length

            if value is None:
                return
            range_name = named_range_name or _convert_to_camel_case(name)
            ps.append([name, value])
            def_range = '$B$%d' % ps.max_row
            defn = DefinedName(range_name, attr_text='%s!%s' % (quote_sheetname(ps.title), def_range))
            wb.defined_names[range_name] = defn
            max_name_length = max(len(name), max_name_length)

        add_param_value('Baseline year', context.instance.reference_year, 'BaselineYear')
        add_param_value('Target year', context.target_year)
        add_param_value('Model end year', context.model_end_year)
        ps.column_dimensions['A'].width = max_name_length

    def _create_workbook(self, wb_contents: BytesIO | None) -> Workbook:
        if wb_contents is None:
            # Create a new workbook
            wb = Workbook()
            del wb[wb.sheetnames[0]]
            return wb

        wb = load_workbook(wb_contents, rich_text=True)
        to_remove: set[str] = set()
        for name, defn in wb.defined_names.items():
            for sheet_name, _ in defn.destinations:
                if sheet_name in self._created_sheet_names:
                    to_remove.add(name)
                    break
        for def_name in to_remove:
            del wb.defined_names[def_name]
        return wb

    def _get_cache_key(self, url: str) -> str:
        return 'excel_url:%s' % hashlib.md5(url.encode('utf-8'), usedforsecurity=False).hexdigest()

    def _get_from_cache(self, url: str) -> BytesIO | None:
        from django.core.cache import cache

        url_hash = self._get_cache_key(url)
        try:
            data = cache.get('excel_url:%s' % url_hash, default=None)
        except Exception as e:
            logger.exception('Error getting from cache: %s' % str(e))
            return None
        if data is None:
            return None
        logger.info('Cache hit for %s' % url)
        return BytesIO(data)

    def _put_in_cache(self, url: str, data: BytesIO) -> None:
        from django.core.cache import cache

        url_hash = self._get_cache_key(url)
        try:
            cache.set('excel_url:%s' % url_hash, data.getvalue(), timeout=7 * 24 * 60 * 60)
        except Exception as e:
            logger.exception('Error putting in cache: %s' % str(e))

    def _download_wb(self, url: str) -> BytesIO:
        logger.info('Downloading results excel from: %s' % url)
        resp = requests.get(url, timeout=(10, 30))  # 5 seconds for connection, 30 seconds for read
        logger.info('File downloaded, status %d' % resp.status_code)
        resp.raise_for_status()
        return BytesIO(resp.content)

    def _create_base(self, existing_wb: Path | str | None = None) -> Workbook:
        if existing_wb is None:
            existing_wb = self.base_excel_url

        if isinstance(existing_wb, Path):
            wb_contents = BytesIO(existing_wb.read_bytes())
        elif isinstance(existing_wb, str):
            cache_data = self._get_from_cache(existing_wb)
            if cache_data is None:
                try:
                    wb_contents = self._download_wb(existing_wb)
                except requests.RequestException as e:
                    logger.error('Unable to download workbook from URL: %s' % str(e))
                    raise
                self._put_in_cache(existing_wb, wb_contents)
            else:
                wb_contents = cache_data
        else:
            wb_contents = None

        wb = self._create_workbook(wb_contents)

        for sn in self._created_sheet_names:
            if sn in wb:
                del wb[sn]
        return wb

    def create_result_excel(  # noqa: C901, PLR0912, PLR0915
        self,
        instance: Instance,
        existing_wb: Path | str | None = None,
    ) -> BytesIO:
        """
        Create or update an Excel workbook with simulation results.

        This function generates an Excel workbook containing simulation results from the provided model (Context).
        It creates or updates sheets for data and parameters, and defines named ranges for easy reference.

        Args:
            instance (Instance): The simulation context instance containing nodes, dimensions, and other data.
            existing_wb (Path | str | None, optional): Path or URL to an existing workbook to update.
                                                    If None, a new workbook is created. Defaults to None.
                                                    If a string is provided, it's treated as a URL.

        Returns:
            BytesIO: A buffer containing the Excel workbook data.

        Notes:
            - The function creates two sheets: 'Data' and 'Parameters'.
            - It defines named ranges for each column in the 'Data' sheet and for parameters.
            - If updating an existing workbook, it removes and recreates the 'Data' and 'Parameters' sheets.
            - The 'Data' sheet includes columns for Node, Year, Unit, Value, BaselineValue, Forecast,
            dimension values, and impact of actions.
            - The 'Parameters' sheet includes simulation parameters like Baseline year and Target year.
            - If a URL is provided for existing_wb, the function will attempt to download the workbook
            from that URL before updating it.
            - If node_ids is provided, only the specified nodes will be included in the output.
            Otherwise, all outcome nodes from the context will be used.

        """

        pc = PerfCounter('generate excel', level=PerfCounter.Level.DEBUG)
        context = instance.context
        if self.node_ids is None:
            node_ids = [node.id for node in context.get_outcome_nodes()]
        else:
            node_ids = self.node_ids

        baseline = context.scenarios[BASELINE_SCENARIO]
        nodes = [context.nodes[node_id] for node_id in node_ids]

        nodes_without_baseline = [node for node in nodes if not node.baseline_values_calculated()]
        with (
            context.start_span('calculate baseline for %d nodes' % len(nodes_without_baseline), op=MODEL_CALC_OP),
            baseline.override(set_active=True),
        ):
            context.prefetch_node_cache(nodes_without_baseline)
            for node in nodes_without_baseline:
                node.get_baseline_values()

        pc.display('baseline calculated')
        all_dims = set()

        for node in nodes:
            for dim in node.output_dimensions.values():
                if dim.id not in all_dims:
                    all_dims.add(dim.id)

        wb = self._create_base(existing_wb)
        ds: Worksheet = wb.create_sheet(DATA_SHEET_NAME)

        dims = [context.dimensions[dim_id] for dim_id in all_dims]
        if self.format == 'wide':
            cols = FIXED_COLUMNS_WIDE + [dim_id for dim_id in all_dims]  # noqa: C416
        else:
            cols = FIXED_COLUMNS_LONG + [_dim_to_col(dim) for dim in dims]

        if self.action_ids is not None:
            actions = [context.get_action(a) for a in self.action_ids]
            aseq = True
        else:
            actions = context.get_actions()[1:]
            aseq = False

        if self.format == 'wide':
            all_years: set[int] = set()
            for node in nodes:
                df = node.get_output_pl()
                all_years.update(df[YEAR_COLUMN].unique())

            cols = cols + [str(year) for year in sorted(all_years)]
        else:
            for act in actions:
                cols.append('Impact_%s' % act.id)

        ds.append(cols)

        for idx, col in enumerate(cols):
            letter = get_column_letter(idx + 1)
            ref = '%s!%s' % (quote_sheetname(ds.title), '$%s:$%s' % (letter, letter))
            defn = DefinedName(col, attr_text=ref)
            wb.defined_names[col] = defn

            cell: Cell = ds['%s1' % letter]
            cell.border = Border(bottom=Side(style='thin', color='222222'))
            cell.font = Font(bold=True)

        dim_ids = [dim.id for dim in dims]
        for node in nodes:
            with context.start_span('output for node: %s' % node.id, op='function'):
                if self.format == 'wide':
                    self._output_node_wide(sheet=ds, node=node, dim_ids=dim_ids, cols=cols)
                else:
                    self._output_node_long(context, wb, ds, node, dim_ids=dim_ids,
                                  actions=actions, aseq=aseq)
        ds.freeze_panes = ds['B2']

        ds.column_dimensions['A'].width = 20
        # for col in range(ds.min_column, ds.max_column + 1):
        #    ds.column_dimensions[get_column_letter(col)].bestFit = True
        #    ds.column_dimensions[get_column_letter(col)].auto_size = True

        self._add_param_sheet(context, wb)

        buffer = BytesIO()
        wb.save(buffer)
        return buffer

    @classmethod
    def create_for_instance(
        cls, ic: InstanceConfig, existing_wb: Path | str | None = None, context: Context | None = None,
        format: str | None = None,
    ) -> BytesIO:
        if context is None:
            instance = ic.get_instance()
            context = instance.context
        else:
            instance = context.instance

        if instance.result_excels:
            excel_res = instance.result_excels[0]  # FIXME
        else:
            if (
                not ic.has_framework_config()
                or not (fw := ic.framework_config.framework).result_excel_url # type: ignore
                or not fw.result_excel_node_ids
            ):
                raise ExportNotSupportedError("Framework '%s' doesn't support generating result files" % instance.id)

            excel_res = InstanceResultExcel(
                name=fw.name,
                base_excel_url=fw.result_excel_url,
                node_ids=fw.result_excel_node_ids,
                format=format,
            )

        with context.start_span('create result excel', op='function'), PerfCounter.time_it() as pc:
            res = excel_res.create_result_excel(instance, existing_wb)
            logger.info('Excel creation took %.1f ms' % pc.finish())
        return res
